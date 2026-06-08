from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import LineChart, Reference
import csv
import sys
import os

if len(sys.argv) != 2:
    print("Uso:")
    print("python excel_cache.py arquivo.csv")
    exit()

csv_file = sys.argv[1]

nome_saida = os.path.splitext(csv_file)[0] + ".xlsx"

wb = Workbook()
ws = wb.active
ws.title = "Resultados"

with open(csv_file, encoding="utf-8") as f:
    reader = csv.reader(f, delimiter=';')

    for linha in reader:
        ws.append(linha)

for cell in ws[1]:
    cell.font = Font(bold=True)

for coluna in ws.columns:
    tamanho = max(len(str(c.value)) if c.value is not None else 0 for c in coluna)
    ws.column_dimensions[coluna[0].column_letter].width = tamanho + 5

ultima_linha = ws.max_row
ultima_coluna = ws.max_column

from openpyxl.utils import get_column_letter

ref = f"A1:{get_column_letter(ultima_coluna)}{ultima_linha}"

tabela = Table(
    displayName="ResultadosCache",
    ref=ref
)

estilo = TableStyleInfo(
    name="TableStyleMedium9",
    showRowStripes=True,
    showColumnStripes=False
)

tabela.tableStyleInfo = estilo
ws.add_table(tabela)

ws.freeze_panes = "A2"

grafico_sheet = wb.create_sheet("Grafico")

chart = LineChart()
chart.title = "Taxa de Acerto"
chart.y_axis.title = "Hit Rate (%)"
chart.x_axis.title = ws.cell(1,1).value

dados = Reference(
    ws,
    min_col=2,
    min_row=1,
    max_row=ultima_linha
)

categorias = Reference(
    ws,
    min_col=1,
    min_row=2,
    max_row=ultima_linha
)

chart.add_data(dados, titles_from_data=True)
chart.set_categories(categorias)

grafico_sheet.add_chart(chart, "A1")

wb.save(nome_saida)

print(f"Arquivo gerado: {nome_saida}")